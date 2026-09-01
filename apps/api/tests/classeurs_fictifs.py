"""Classeurs fabriqués pour les tests : hostiles, tordus, ou simplement normaux.

Fabriqués et non commités, pour la même raison que les images de `logo` : un
`.xlsx` est une archive binaire, et un binaire commité est un bloc que personne
ne relit. Ici, on voit exactement ce que chaque cas porte.
"""

from __future__ import annotations

import io
import zipfile
from typing import Any

ENTETES = ["code", "libelle", "unite", "prix_unitaire"]
LIGNE = ["MAT-FIC-001", "Sable fictif 0/4", "m3", 41.5]


def classeur(
    lignes: list[list[Any]] | None = None,
    *,
    entetes: list[str] | None = None,
    feuilles: dict[str, list[list[Any]]] | None = None,
    nom_de_feuille: str = "Prix",
) -> bytes:
    """Un classeur ordinaire, ou plusieurs feuilles si on le demande."""
    from openpyxl import Workbook

    livre = Workbook()
    if feuilles:
        premier = True
        for nom, contenu in feuilles.items():
            onglet = livre.active if premier else livre.create_sheet()
            onglet.title = nom
            for ligne in contenu:
                onglet.append(ligne)
            premier = False
    else:
        onglet = livre.active
        onglet.title = nom_de_feuille
        onglet.append(entetes if entetes is not None else ENTETES)
        for ligne in lignes if lignes is not None else [LIGNE]:
            onglet.append(ligne)

    tampon = io.BytesIO()
    livre.save(tampon)
    return tampon.getvalue()


def avec_une_formule() -> bytes:
    """Un classeur dont le prix est calculé plutôt qu'écrit."""
    from openpyxl import Workbook

    livre = Workbook()
    onglet = livre.active
    onglet.title = "Prix"
    onglet.append(ENTETES)
    onglet.append(["MAT-FIC-001", "Sable fictif 0/4", "m3", "=34.30*1.21"])
    tampon = io.BytesIO()
    livre.save(tampon)
    return tampon.getvalue()


def reecrire(octets: bytes, ajouts: dict[str, bytes]) -> bytes:
    """Recopie une archive en y ajoutant des entrées.

    Passer par une vraie archive plutôt que d'en fabriquer une de toutes pièces
    garantit que seul l'AJOUT distingue le cas du cas nominal : si le test
    échouait pour une autre raison, il le ferait aussi sur le classeur ordinaire.
    """
    tampon = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(octets)) as source,
        zipfile.ZipFile(tampon, "w", zipfile.ZIP_DEFLATED) as cible,
    ):
        for entree in source.infolist():
            cible.writestr(entree, source.read(entree.filename))
        for nom, contenu in ajouts.items():
            cible.writestr(nom, contenu)
    return tampon.getvalue()


def avec_des_macros() -> bytes:
    return reecrire(classeur(), {"xl/vbaProject.bin": b"\x00faux objet VBA"})


def avec_un_lien_externe() -> bytes:
    return reecrire(
        classeur(),
        {"xl/externalLinks/externalLink1.xml": b"<externalLink/>"},
    )


def avec_un_chemin_remontant() -> bytes:
    return reecrire(classeur(), {"../evasion.xml": b"<x/>"})


def bombe_de_decompression(*, entrees: int = 8, par_entree: int = 12 * 1024 * 1024) -> bytes:
    """Une archive minuscule qui se déplie au-delà de la borne.

    Des zéros : DEFLATE les comprime à un millième, si bien que ce fichier tient
    dans quelques kilooctets et annonce près de cent mégaoctets dépliés.
    """
    tampon = io.BytesIO()
    with zipfile.ZipFile(tampon, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.xml", b"<workbook/>")
        for rang in range(entrees):
            archive.writestr(f"xl/media/gros{rang}.bin", b"\x00" * par_entree)
    return tampon.getvalue()


def archive_sans_classeur() -> bytes:
    """Une archive ZIP valide qui n'est pas un classeur."""
    tampon = io.BytesIO()
    with zipfile.ZipFile(tampon, "w") as archive:
        archive.writestr("lisezmoi.txt", b"ceci n'est pas un classeur")
    return tampon.getvalue()


#: Un vieux `.xls`, reconnu à sa signature OLE2 — et un OOXML chiffré aussi.
OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512

#: Une archive tronquée : la signature promet un ZIP que le reste ne tient pas.
ZIP_TRONQUE = b"PK\x03\x04" + b"\x00" * 40
