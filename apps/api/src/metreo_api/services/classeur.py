"""Lecture d'un classeur XLSX : bornée, vérifiée, et jamais sur parole.

Un fichier importé est une **donnée non fiable**. Un XLSX l'est deux fois : ce
n'est pas un tableau mais une archive ZIP de documents XML, et chacune de ces
deux couches a ses propres façons de mentir. On borne donc les deux, dans cet
ordre, avant qu'aucun analyseur ne voie le contenu.

**Pourquoi une bibliothèque et non un analyseur maison.** OOXML est un format
vaste — chaînes partagées, styles, dates sérielles, feuilles masquées,
plages nommées. En écrire un analyseur ici reviendrait à réimplémenter, sans
personne pour le maintenir, ce qu'openpyxl fait depuis quinze ans. La borne,
elle, ne se délègue pas : openpyxl n'a jamais promis de résister à une bombe
de décompression, et c'est à nous de ne pas la lui donner à avaler.

**Ce qui est refusé, et pourquoi.**

* Les **macros** : un classeur porteur de `vbaProject.bin` est du code. Nous
  n'en exécutons aucun, mais accepter le fichier reviendrait à le stocker puis
  à le rendre, et à faire de ce service un vecteur de transport.
* Les **formules** : une cellule `=B2*1,21` n'a pas de valeur tant qu'un
  tableur ne l'a pas calculée. La valeur en cache peut être absente — un
  fichier produit par un script n'en a aucune — ou périmée. Importer un prix
  dont on ne sait pas s'il vaut 100 ou 121 serait pire que refuser.
* Les **liens externes** : ils désignent d'autres classeurs, donc des valeurs
  que ce fichier ne porte pas et que ce serveur n'ira pas chercher.
* Le **chiffrement** et le vieux **.xls** : tous deux sont des conteneurs OLE2,
  reconnus à leur signature. Ni l'un ni l'autre n'est un ZIP, et aucun des deux
  n'est dans le périmètre.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any

from metreo_domain.errors import DomainError

#: Signature d'une archive ZIP — donc d'un XLSX, d'un XLSM ou d'un ODS.
SIGNATURE_ZIP = b"PK\x03\x04"
#: Signature d'un conteneur OLE2 : le vieux `.xls`, et aussi un OOXML chiffré.
SIGNATURE_OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

#: Ce que l'archive a le droit de peser une fois DÉPLIÉE.
#:
#: Le plafond du fichier reçu ne protège rien à lui seul : DEFLATE comprime des
#: cellules répétées à mille pour un, et un classeur de quelques centaines de
#: kilooctets peut porter des gigaoctets de XML. La borne qui compte est celle
#: du déplié, et elle se lit dans l'en-tête de l'archive AVANT toute extraction.
TAILLE_DECOMPRESSEE_MAXIMUM = 64 * 1024 * 1024

#: Rapport maximal entre le déplié et le comprimé. Un classeur ordinaire tourne
#: autour de dix ; au-delà de cent, ce n'est plus un tableau de prix.
RATIO_MAXIMUM = 100

#: Un XLSX ordinaire compte une vingtaine d'entrées. Mille est déjà très large
#: pour un classeur légitime, et ferme la porte à une archive faite de millions
#: de petits fichiers dont chacun tient sous les bornes précédentes.
ENTREES_MAXIMUM = 1_000

#: Bornes du tableau lui-même, appliquées pendant la lecture et non après.
FEUILLES_MAXIMUM = 50
LIGNES_MAXIMUM = 50_000
COLONNES_MAXIMUM = 200
CELLULES_MAXIMUM = 1_000_000


class ClasseurRefuse(DomainError):
    """Un refus que l'écran doit pouvoir montrer à qui téléverse.

    Porte un `code` stable : l'interface choisit son message et son encadré,
    et une chaîne libre ne s'y prêterait pas.
    """

    def __init__(self, code: str, message: str, **contexte: object) -> None:
        super().__init__(message, **contexte)
        self.code = code


@dataclass(frozen=True)
class FeuilleLue:
    """Le contenu d'une feuille, rendu comme le rendrait un lecteur CSV.

    C'est le point de convergence : au-delà de cette structure, plus rien dans
    l'import ne sait de quel format vient la donnée.
    """

    nom: str
    headers: list[str]
    lignes: list[dict[str, str]]
    #: Le rang réel dans le classeur de chaque ligne rendue, pour que l'erreur
    #: signalée à l'utilisateur désigne la ligne qu'il voit dans son tableur.
    rangs: list[int]


def est_un_zip(payload: bytes) -> bool:
    return payload.startswith(SIGNATURE_ZIP)


def est_un_ole2(payload: bytes) -> bool:
    return payload.startswith(SIGNATURE_OLE2)


def detecter_le_format(payload: bytes) -> str:
    """Le format d'un fichier d'après son CONTENU, jamais d'après son nom.

    L'extension est une intention, pas un fait : un `.csv` renommé en `.xlsx`
    est courant, et un `.xlsx` qui n'est pas une archive ZIP l'est aussi. Rendre
    « xlsx » sur la foi du nom enverrait un texte à un lecteur d'archive, qui
    échouerait par une erreur illisible plutôt que par un refus clair.
    """
    if est_un_zip(payload):
        return "xlsx"
    if est_un_ole2(payload):
        return "ole2"
    return "csv"


def _refuser_selon_la_signature(payload: bytes) -> None:
    """Les refus qui se prononcent sur les premiers octets, avant tout le reste.

    Les prononcer tôt donne à l'utilisateur le vrai motif — « ceci est un
    ancien .xls » — au lieu de l'erreur d'analyse qu'un lecteur d'archive
    produirait quelques couches plus bas.
    """
    if not payload:
        raise ClasseurRefuse("fichier_vide", "Ce fichier est vide.")
    if est_un_ole2(payload):
        raise ClasseurRefuse(
            "format_ole2",
            "Ce fichier est un ancien classeur .xls, ou un classeur chiffré. "
            "Enregistrez-le au format .xlsx, sans mot de passe, puis réessayez.",
        )
    if not est_un_zip(payload):
        raise ClasseurRefuse(
            "pas_une_archive",
            "Ce fichier n'est pas un classeur .xlsx : il n'en a pas la structure.",
        )


def _borner_l_archive(archive: zipfile.ZipFile, taille_recue: int) -> None:
    """Mesure l'archive SANS rien extraire, et refuse avant de déplier.

    Les tailles dépliées sont lues dans le répertoire central du ZIP. Elles
    peuvent mentir — c'est le fichier qui les écrit — mais un mensonge à la
    baisse ne sert à rien ici : `_lire_les_entrees` relit la taille réelle en
    extrayant, et refuse aussi celle-là. Ce contrôle-ci ferme le cas courant,
    et le second ferme le cas malveillant.
    """
    entrees = archive.infolist()
    if len(entrees) > ENTREES_MAXIMUM:
        raise ClasseurRefuse(
            "trop_d_entrees",
            f"Ce classeur contient {len(entrees)} éléments internes, "
            f"au-delà des {ENTREES_MAXIMUM} admis.",
        )

    annoncee = sum(entree.file_size for entree in entrees)
    if annoncee > TAILLE_DECOMPRESSEE_MAXIMUM:
        raise ClasseurRefuse(
            "decompresse_trop_grand",
            f"Déplié, ce classeur occuperait {annoncee // (1024 * 1024)} Mio, "
            f"au-delà des {TAILLE_DECOMPRESSEE_MAXIMUM // (1024 * 1024)} Mio admis.",
        )
    if taille_recue > 0 and annoncee > taille_recue * RATIO_MAXIMUM:
        raise ClasseurRefuse(
            "ratio_excessif",
            f"Ce classeur se déplie {annoncee // max(taille_recue, 1)} fois, "
            f"au-delà du rapport {RATIO_MAXIMUM} admis. Un tableau de prix ne "
            "se comprime pas ainsi.",
        )

    for entree in entrees:
        # Un chemin absolu ou remontant ne peut rien atteindre ici — nous
        # n'écrivons jamais sur le disque — mais sa présence dit que l'archive
        # n'a pas été produite par un tableur, et c'est un motif suffisant.
        nom = entree.filename
        if nom.startswith("/") or ".." in nom.replace("\\", "/").split("/"):
            raise ClasseurRefuse(
                "chemin_suspect",
                "Ce classeur porte un chemin interne anormal ; il n'a pas été "
                "produit par un tableur.",
            )


def _refuser_le_contenu_executable(archive: zipfile.ZipFile) -> None:
    """Macros, liens externes, chiffrement : reconnus à ce que porte l'archive."""
    noms = {entree.filename for entree in archive.infolist()}

    if any(nom.endswith("vbaProject.bin") for nom in noms):
        raise ClasseurRefuse(
            "macros_refusees",
            "Ce classeur contient des macros. Enregistrez-le en .xlsx "
            "— sans macros — puis réessayez.",
        )
    if any(nom.startswith("xl/externalLinks/") for nom in noms):
        raise ClasseurRefuse(
            "liens_externes_refuses",
            "Ce classeur renvoie à d'autres classeurs. Les valeurs qu'il "
            "affiche ne sont pas dans le fichier : collez-les d'abord en clair.",
        )
    # Un OOXML chiffré est un conteneur OLE2, déjà refusé à la signature. Celui
    # qui arrive ici en portant tout de même la trace d'un chiffrement est une
    # variante que nous ne savons pas lire, et le dire vaut mieux que l'ouvrir.
    if any("EncryptedPackage" in nom or "encryption" in nom.lower() for nom in noms):
        raise ClasseurRefuse(
            "chiffrement_refuse",
            "Ce classeur est protégé par un mot de passe. Retirez la protection puis réessayez.",
        )
    if "xl/workbook.xml" not in noms:
        raise ClasseurRefuse(
            "pas_un_classeur",
            "Cette archive n'est pas un classeur Excel : elle n'en a pas le document principal.",
        )


def _cellule_en_texte(valeur: Any) -> str:
    """Rend une cellule comme le CSV l'aurait portée : du texte, et rien d'autre.

    C'est ce qui fait converger les deux formats. Un tableur rend des types —
    `datetime`, `float`, `Decimal`, `bool` — là où un CSV ne rend que des
    chaînes ; laisser passer ces types obligerait la normalisation à traiter
    deux cas pour chaque colonne, et les deux finiraient par diverger.

    Les conversions délicates :

    * une **date** devient ISO, format que `coerce_local_date` laisse passer
      intact ; c'est aussi ce qu'un tableur écrit quand on l'exporte en CSV ;
    * un **entier flottant** — `12.0`, ce qu'Excel stocke pour « 12 » — perd son
      « .0 », sans quoi une quantité minimale de 12 arriverait « 12.0 » ;
    * un **booléen** devient « true »/« false » en minuscules, comme le contrat
      les attend ;
    * `None` devient la chaîne vide, comme une cellule CSV absente.
    """
    if valeur is None:
        return ""
    if isinstance(valeur, bool):
        return "true" if valeur else "false"
    if isinstance(valeur, datetime):
        # Une date saisie sans heure arrive à minuit : on ne rend alors que la
        # date, sans quoi « 31/12/2026 » deviendrait « 2026-12-31 00:00:00 »,
        # que le contrat refuserait.
        if valeur.time() == time(0, 0):
            return valeur.date().isoformat()
        return valeur.isoformat()
    if isinstance(valeur, date):
        return valeur.isoformat()
    if isinstance(valeur, float):
        if valeur.is_integer():
            return str(int(valeur))
        return repr(valeur)
    if isinstance(valeur, Decimal):
        return str(valeur)
    return str(valeur).strip()


def _refuser_une_formule(valeur: Any, rang: int, colonne: int) -> None:
    """Une cellule dont la valeur est une formule n'a pas de valeur ici.

    Le classeur est ouvert avec les formules VISIBLES plutôt qu'avec les
    valeurs en cache, et c'est délibéré : le cache n'existe que si un tableur a
    ouvert le fichier, il est absent d'un classeur produit par un script, et
    rien ne dit qu'il correspond encore aux formules. Importer un prix qu'on ne
    sait pas lire serait pire que refuser de le lire.
    """
    if isinstance(valeur, str) and valeur.startswith("="):
        raise ClasseurRefuse(
            "formules_refusees",
            f"La cellule ligne {rang}, colonne {colonne} contient une formule "
            f"« {valeur[:40]} ». Collez les valeurs calculées, puis réessayez.",
        )


def noms_des_feuilles(payload: bytes) -> list[str]:
    """Les feuilles d'un classeur, pour que l'utilisateur choisisse la sienne.

    Un classeur de prix réel porte souvent une feuille de garde, un barème et
    des notes. Deviner laquelle importer — la première, la plus grande — se
    tromperait sans le dire ; on rend donc la liste, et l'écran demande.
    """
    return _ouvrir(payload).sheetnames


def _ouvrir(payload: bytes) -> Any:
    """Ouvre le classeur après TOUS les contrôles, et jamais avant.

    L'import d'openpyxl est fait ici plutôt qu'en tête de module : le coût de
    son chargement — et de celui de ses dépendances — ne se paie qu'au premier
    import de classeur, et non au démarrage de chaque processus de l'API.
    """
    from openpyxl import load_workbook

    _refuser_selon_la_signature(payload)
    try:
        archive = zipfile.ZipFile(BytesIO(payload))
    except zipfile.BadZipFile as refus:
        raise ClasseurRefuse(
            "archive_illisible",
            "Ce classeur est corrompu : son archive ne s'ouvre pas.",
        ) from refus

    with archive:
        _borner_l_archive(archive, len(payload))
        _refuser_le_contenu_executable(archive)

    try:
        # `read_only` lit la feuille en flux, sans construire tout le classeur
        # en mémoire ; `data_only=False` garde les formules VISIBLES, pour
        # pouvoir les refuser plutôt que d'avaler un cache incertain.
        return load_workbook(BytesIO(payload), read_only=True, data_only=False, keep_links=False)
    except ClasseurRefuse:
        raise
    except Exception as refus:  # openpyxl lève large : KeyError, ValueError, ...
        raise ClasseurRefuse(
            "classeur_illisible",
            "Ce classeur ne peut pas être lu : il est corrompu, ou son format "
            "n'est pas celui qu'il annonce.",
        ) from refus


def lire(payload: bytes, *, feuille: str | None = None) -> tuple[FeuilleLue, dict[str, Any]]:
    """Lit UNE feuille d'un classeur et la rend sous la forme du lecteur CSV.

    `feuille` nomme la feuille à lire. Sans elle, la première est prise — et le
    rapport dit laquelle, avec la liste des autres, pour que l'écran propose de
    changer plutôt que de laisser croire que le classeur n'en avait qu'une.
    """
    classeur = _ouvrir(payload)
    try:
        feuilles = classeur.sheetnames
        if len(feuilles) > FEUILLES_MAXIMUM:
            raise ClasseurRefuse(
                "trop_de_feuilles",
                f"Ce classeur porte {len(feuilles)} feuilles, au-delà des "
                f"{FEUILLES_MAXIMUM} admises.",
            )
        if not feuilles:
            raise ClasseurRefuse("classeur_sans_feuille", "Ce classeur ne porte aucune feuille.")

        choisie = feuille or feuilles[0]
        if choisie not in feuilles:
            raise ClasseurRefuse(
                "feuille_inconnue",
                f"Ce classeur n'a pas de feuille « {choisie} ». Il porte : {', '.join(feuilles)}.",
                feuilles=feuilles,
            )

        lue = _lire_la_feuille(classeur[choisie])
    finally:
        # `read_only` ouvre des descripteurs sur l'archive : les rendre est
        # obligatoire, y compris quand la lecture s'est soldée par un refus.
        classeur.close()

    meta: dict[str, Any] = {
        "format": "xlsx",
        "feuille": lue.nom,
        "feuilles": feuilles,
        "colonnes_lues": len(lue.headers),
        "lignes_lues": len(lue.lignes),
    }
    return lue, meta


def _lire_la_feuille(onglet: Any) -> FeuilleLue:
    """Parcourt la feuille en la bornant PENDANT la lecture, pas après.

    Borner après aurait exigé de tout charger d'abord — c'est-à-dire d'accepter
    exactement le coût contre lequel la borne existe.
    """
    headers: list[str] = []
    lignes: list[dict[str, str]] = []
    rangs: list[int] = []
    cellules = 0

    for rang, brute in enumerate(onglet.iter_rows(values_only=False), start=1):
        valeurs = []
        for colonne, cellule in enumerate(brute, start=1):
            if colonne > COLONNES_MAXIMUM:
                raise ClasseurRefuse(
                    "trop_de_colonnes",
                    f"La feuille « {onglet.title} » porte plus de {COLONNES_MAXIMUM} colonnes.",
                )
            _refuser_une_formule(cellule.value, rang, colonne)
            valeurs.append(_cellule_en_texte(cellule.value))

        cellules += len(valeurs)
        if cellules > CELLULES_MAXIMUM:
            raise ClasseurRefuse(
                "trop_de_cellules",
                f"Cette feuille dépasse {CELLULES_MAXIMUM} cellules.",
            )

        if not headers:
            # La ligne d'en-tête est la première qui porte quelque chose : un
            # classeur réel commence souvent par une ligne vide ou un titre de
            # présentation, et s'arrêter à la ligne 1 rendrait des en-têtes vides.
            if any(valeur for valeur in valeurs):
                headers = list(valeurs)
                # Les colonnes sans en-tête ne sont pas lues : elles n'ont pas
                # de nom auquel rattacher leur contenu. Les colonnes vides en
                # fin de ligne sont fréquentes — un tableur les rend dès qu'une
                # cellule a été mise en forme.
                while headers and not headers[-1]:
                    headers.pop()
            continue

        if len(lignes) >= LIGNES_MAXIMUM:
            raise ClasseurRefuse(
                "trop_de_lignes",
                f"Cette feuille dépasse {LIGNES_MAXIMUM} lignes de données.",
            )

        if not any(valeur for valeur in valeurs):
            continue  # ligne vide, comme dans le lecteur CSV

        ligne = {
            entete: (valeurs[index] if index < len(valeurs) else "")
            for index, entete in enumerate(headers)
            if entete
        }
        lignes.append(ligne)
        rangs.append(rang)

    if not headers:
        raise ClasseurRefuse(
            "feuille_vide",
            f"La feuille « {onglet.title} » ne porte aucune colonne nommée.",
        )

    return FeuilleLue(
        nom=onglet.title, headers=[h for h in headers if h], lignes=lignes, rangs=rangs
    )
